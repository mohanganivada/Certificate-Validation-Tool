import openpyxl
from folderData import matched_pdf_files, matched_jpg_files
from validation import final_result, final_data
new = []
my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
my_sheet.title = "Results"
my_sheet = my_wb.active
c = my_sheet.cell(row=1, column=1)
c.value = "Certification Name"
c = my_sheet.cell(row=1, column=2)
c.value = "Actual Certification Name"
c = my_sheet.cell(row=1, column=3)
c.value = "Employee Code"
c = my_sheet.cell(row=1, column=4)
c.value = "Employee Name"
c = my_sheet.cell(row=1, column=5)
c.value = "Status"
c = my_sheet.cell(row=1, column=6)
c.value = "Exam Pass Date"
c = my_sheet.cell(row=1, column=7)
c.value = "Reason"
c = my_sheet.cell(row=1, column=8)
c.value = "Certificate Name"
for i in matched_pdf_files:
    new.append(i)
for i in matched_jpg_files:
    new.append(i)
for i, items in enumerate(new):
    c = my_sheet.cell(row=i+2, column=1)
    c.value = str(final_data[items]["Certification Name"])
    c = my_sheet.cell(row=i+2, column=2)
    c.value = str(final_result[items]["Actual Certificate Name"])
    c = my_sheet.cell(row=i+2, column=3)
    c.value = final_data[items]["Employee Code"]
    c = my_sheet.cell(row=i+2, column=4)
    c.value = final_data[items]["Name"]
    c = my_sheet.cell(row=i+2, column=5)
    c.value = final_result[items]['Status']
    date = str(final_data[items]["Exam Pass Date"])
    date = date.split(" ")[0]
    c = my_sheet.cell(row=i+2, column=6)
    c.value = date
    c = my_sheet.cell(row=i+2, column=7)
    c.value = final_result[items]['Reason']
    c = my_sheet.cell(row=i+2, column=8)
    c.value = f"{items}"